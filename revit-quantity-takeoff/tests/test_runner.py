"""runner.py / excel.py 통합 테스트."""
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

from takeoff import runner
from takeoff.models import Member

try:
    import openpyxl  # noqa: F401
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


SAMPLE = [
    {"category": "Columns", "mark": "C1", "type_name": "RC400x600",
     "count": 1, "b": 0.4, "h": 0.6, "height": 3.0},
    {"category": "Columns", "mark": "C1", "type_name": "RC400x600",
     "count": 1, "b": 0.4, "h": 0.6, "height": 3.0},
    {"category": "Foundations", "mark": "F1", "count": 2,
     "b": 2.0, "l": 2.0, "h": 0.5},
]


class GroupingTests(unittest.TestCase):
    def test_identical_members_are_merged(self):
        members = [Member.from_dict(d) for d in SAMPLE]
        grouped = runner.group_members(members)
        # C1 두 개 -> 하나(count=2), F1 하나(count=2)
        self.assertEqual(len(grouped), 2)
        c1 = next(g for g in grouped if g.mark == "C1")
        self.assertEqual(c1.count, 2)

    def test_build_line_items_aggregates_count(self):
        items = runner.build_line_items(SAMPLE)
        c1_concrete = next(i for i in items if i.mark == "C1" and i.quantity == "concrete")
        # 0.4*0.6*3.0*2 = 1.44
        self.assertAlmostEqual(c1_concrete.value, 1.44)
        self.assertEqual(c1_concrete.formula, "0.4×0.6×3×2")

    def test_flat_dims_via_from_dict(self):
        m = Member.from_dict({"category": "Walls", "length": 5.0, "height": 3.0, "thickness": 0.2})
        self.assertEqual(m.dims, {"length": 5.0, "height": 3.0, "thickness": 0.2})


@unittest.skipUnless(HAS_OPENPYXL, "openpyxl 미설치 — 엑셀 생성 테스트 건너뜀")
class ExcelTests(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.out = str(Path(self._tmp.name) / "takeoff.xlsx")

    def tearDown(self):
        self._tmp.cleanup()

    def test_run_creates_workbook_with_expected_totals(self):
        result = runner.run(SAMPLE, self.out, project_info={"공사명": "테스트현장"})
        self.assertTrue(Path(self.out).exists())
        # 콘크리트 합계 = C1(0.4*0.6*3*2=1.44) + F1(2*2*0.5*2=4.0) = 5.44
        self.assertAlmostEqual(result["totals"]["concrete"], 5.44)
        self.assertEqual(result["member_count"], 4)

    def test_workbook_sheets_and_values(self):
        runner.run(SAMPLE, self.out)
        import openpyxl
        wb = openpyxl.load_workbook(self.out)
        self.assertIn("집계", wb.sheetnames)
        self.assertIn("콘크리트 (m³)", wb.sheetnames)
        ws = wb["콘크리트 (m³)"]
        # 헤더 확인
        self.assertEqual(ws.cell(row=1, column=1).value, "No")
        self.assertEqual(ws.cell(row=1, column=6).value, "산출식")
        # 데이터가 한 줄 이상 존재
        self.assertGreater(ws.max_row, 2)

    def test_empty_input_raises(self):
        with self.assertRaises(ValueError):
            runner.run([], self.out)


if __name__ == "__main__":
    unittest.main()
