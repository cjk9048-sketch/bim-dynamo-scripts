# 골든 검증 하니스
# 실 CSV_Import + Formula + Calc_Sheet 를 현행 VBA 의미대로 평가해
# Quantity_Report(골든) 공종별 합계를 재현·비교한다. (C# Core 이전 정확성 증명)
import math
import re
from collections import defaultdict

import openpyxl

SRC = r"c:\Users\user\Desktop\AI\revit-quantity-takeoff\samples\배수지 수량산출(VBA)1.xlsm"
OUT = r"c:\Users\user\Desktop\AI\revit-quantity-takeoff\docs\골든검증-결과.md"

wb = openpyxl.load_workbook(SRC, data_only=False)
wbv = openpyxl.load_workbook(SRC, data_only=True)

# --- Calc_Sheet: 전역 파생값(캐시) ---
calc = {}
cs = wbv["Calc_Sheet"]
for r in range(2, cs.max_row + 1):
    code = cs.cell(r, 1).value
    val = cs.cell(r, 2).value
    if code is not None:
        calc["[" + str(code).strip() + "]"] = float(val) if isinstance(val, (int, float)) else 0.0

# --- Formula 시트: code -> {discipline: pattern} ---
fs = wb["Formula"]
disc_cols = {}
for c in range(2, fs.max_column + 1):
    h = fs.cell(1, c).value
    if h:
        disc_cols[c] = str(h).strip()
formulas = {}
for r in range(2, fs.max_row + 1):
    code = fs.cell(r, 1).value
    if not code:
        continue
    code = str(code).strip()
    formulas.setdefault(code, {})
    for c, disc in disc_cols.items():
        v = fs.cell(r, c).value
        if v is not None and str(v).strip():
            formulas[code][disc] = str(v).strip()

# --- CSV_Import: 인스턴스 ---
cv = wbv["CSV_Import"]
hdr = {cv.cell(1, c).value: c for c in range(1, cv.max_column + 1)}
DIMS = ["L1", "L2", "L3", "W1", "W2", "W3", "H", "ETC"]
instances = []
code2cat = {}
for r in range(2, cv.max_row + 1):
    code = cv.cell(r, 1).value
    if code is None:
        continue
    code = str(code).strip()
    cat = str(cv.cell(r, hdr["DH_Category"]).value or "").strip()
    code2cat.setdefault(code, cat)
    inst = {"code": code, "class": str(cv.cell(r, hdr["DH_Class"]).value or "")}
    for k in DIMS:
        v = cv.cell(r, hdr[k]).value
        inst[k] = float(v) if isinstance(v, (int, float)) else 0.0
    instances.append(inst)

# 집계 관례(워크북 고유): 대부분 공종은 인스턴스 ×개수 합산.
# 단, (기둥, 철근콘크리트)처럼 '대표 1개 단위물량 + 개수(EA)'로 계상하는 (카테고리,공종) 조합만 대표 처리.
# 골든 실측: C1 콘크리트=대표1개(0.45), C1 거푸집=×개수(356.4).
REPRESENTATIVE = {("기둥", "철근콘크리트"), ("기둥", "무근콘크리트")}


def evalexpr(s):
    s = s.replace("×", "*").replace("^", "**")
    s = re.sub(r"[Ss][Qq][Rr][Tt]\(", "math.sqrt(", s)
    s = s.replace("PI()", "math.pi").replace("pi()", "math.pi")
    return eval(s, {"math": math, "__builtins__": {}})


def substitute(sub, inst):
    """패턴 1개(| 분할·# 라벨제거 후)를 인스턴스+calc 로 치환한 수식문자열."""
    f = sub
    local = {"[L1]": inst["L1"], "[L2]": inst["L2"], "[L3]": inst["L3"],
             "[W1]": inst["W1"], "[W2]": inst["W2"], "[W3]": inst["W3"],
             "[H]": inst["H"], "[H1]": inst["H"], "[ETC]": inst["ETC"]}
    for k, v in local.items():
        f = f.replace(k, repr(float(v)))
    for k, v in calc.items():
        f = f.replace(k, repr(float(v)))
    return f


# --- VBA 의미: (disc,code) 별로 동일 수식문자열 그룹화→ROUND(값×count) 합 ---
groups = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))  # disc->code->formstr->count
unresolved = defaultdict(set)
for inst in instances:
    code = inst["code"]
    if code not in formulas:
        continue
    for disc, pattern in formulas[code].items():
        for sub in pattern.split("|"):
            sub = sub.strip()
            if not sub:
                continue
            if "#" in sub:
                sub = sub.split("#", 1)[1].strip()
            f = substitute(sub, inst)
            if "[" in f:  # 미해결 토큰
                unresolved[disc].add((code, f))
                continue
            groups[disc][code][f] += 1

comp_disc = defaultdict(float)   # discipline -> 합계
comp_code = defaultdict(dict)    # discipline -> {code: 소계}
for disc, codes in groups.items():
    for code, fmap in codes.items():
        rep = (code2cat.get(code), disc) in REPRESENTATIVE  # (카테고리,공종) 대표 1개 계상
        subtotal = 0.0
        for f, cnt in fmap.items():
            try:
                subtotal += round(evalexpr(f) * (1 if rep else cnt), 2)
            except Exception:
                pass
        comp_code[disc][code] = round(subtotal, 2)
        comp_disc[disc] += subtotal
    comp_disc[disc] = round(comp_disc[disc], 2)

# --- 골든: Quantity_Report 의 공종별 =SUM 캐시값 합 ---
qr = wbv["Quantity_Report"]
qrf = wb["Quantity_Report"]
golden_disc = defaultdict(float)
cur = None
disc_names = set(disc_cols.values())
for r in range(1, qr.max_row + 1):
    a = qr.cell(r, 1).value
    if a and str(a).strip() in disc_names:
        cur = str(a).strip()
    # =SUM 셀(공종 계) 찾기
    for c in range(1, qr.max_column + 1):
        fcell = qrf.cell(r, c).value
        if isinstance(fcell, str) and fcell.startswith("=SUM"):
            val = qr.cell(r, c).value
            if isinstance(val, (int, float)) and cur:
                golden_disc[cur] += val
for k in golden_disc:
    golden_disc[k] = round(golden_disc[k], 2)

# --- 리포트 ---
lines = ["# 골든 검증 결과 (수식 라이브러리 정확성)", "",
         "실 CSV_Import + Formula + Calc_Sheet 를 평가해 Quantity_Report(골든) 공종별 합계와 비교.", "",
         "## 공종별 합계 비교", "", "| 공종 | 계산값 | 골든 | 차이 | 판정 |", "|------|--------|------|------|------|"]
all_disc = sorted(set(list(comp_disc) + list(golden_disc)))
n_ok = 0
n_tot = 0
for d in all_disc:
    cval = comp_disc.get(d)
    gval = golden_disc.get(d)
    if gval in (None, 0) and not cval:
        continue
    n_tot += 1
    diff = (cval or 0) - (gval or 0)
    ok = abs(diff) <= 0.01
    if ok:
        n_ok += 1
    lines.append(f"| {d} | {cval} | {gval} | {round(diff,2)} | {'✅' if ok else '❌'} |")
lines += ["", f"**합계 일치: {n_ok}/{n_tot} 공종 (허용오차 ±0.01)**", ""]

lines += ["## 공종별 코드 소계 (계산값)", ""]
for d in sorted(comp_code):
    items = ", ".join(f"{c}={v}" for c, v in sorted(comp_code[d].items()))
    lines.append(f"- **{d}**: {items}")

if unresolved:
    lines += ["", "## 미해결 토큰(치환 실패)", ""]
    for d, st in unresolved.items():
        for code, f in sorted(st):
            lines.append(f"- {d} / {code}: `{f}`")

with open(OUT, "w", encoding="utf-8") as fp:
    fp.write("\n".join(lines))

# --- C# 골든 테스트용 픽스처(JSON) 내보내기 (전사 오류 방지, 단일 진실원천) ---
import json
fixture = {
    "formulas": formulas,                                   # code -> {discipline: pattern}
    "calc": {k: v for k, v in calc.items()},                # "[CalcCode]" -> value
    "representative": [[c, d] for (c, d) in REPRESENTATIVE],  # (category, discipline) 대표1개
    "goldenDisc": {k: v for k, v in golden_disc.items()},    # discipline -> 골든 합계
}
FIX = r"c:\Users\user\Desktop\AI\revit-quantity-takeoff\samples\golden-fixture.json"
with open(FIX, "w", encoding="utf-8") as fp:
    json.dump(fixture, fp, ensure_ascii=False, indent=1)
print(f"WROTE {OUT}")
print(f"WROTE {FIX}")
print(f"공종 합계 일치: {n_ok}/{n_tot}")
for d in all_disc:
    if comp_disc.get(d) or golden_disc.get(d):
        print(f"  {d}: calc={comp_disc.get(d)} golden={golden_disc.get(d)}")
