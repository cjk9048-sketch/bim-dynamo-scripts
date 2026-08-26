# 현행 VBA 워크북에서 시트 구조/수식 추출 (분석용)
import openpyxl

SRC = r"c:\Users\user\Desktop\AI\revit-quantity-takeoff\samples\배수지 수량산출(VBA)1.xlsm"
OUT = r"c:\Users\user\Desktop\AI\revit-quantity-takeoff\docs\현행-VBA-워크북-추출.md"

wb = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
wbv = openpyxl.load_workbook(SRC, data_only=True)  # 캐시된 값

lines = []
def w(s=""):
    lines.append(str(s))

w("# 현행 VBA 워크북 추출")
w()
w("## 시트 목록")
for ws in wb.worksheets:
    w(f"- **{ws.title}** : dims={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")
w()

def dump_sheet(name, max_r=60, max_c=12, with_value=False):
    if name not in wb.sheetnames:
        w(f"## {name} — (없음)")
        return
    ws = wb[name]
    wsv = wbv[name] if name in wbv.sheetnames else None
    w(f"## {name}  (max_row={ws.max_row}, max_col={ws.max_column})")
    r_end = min(ws.max_row, max_r)
    c_end = min(ws.max_column, max_c)
    for r in range(1, r_end + 1):
        cells = []
        for c in range(1, c_end + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None or v == "":
                continue
            col = cell.column_letter
            txt = str(v)
            if with_value and wsv is not None:
                vv = wsv.cell(row=r, column=c).value
                if vv is not None and str(vv) != txt:
                    txt = f"{txt}  ⟶[{vv}]"
            cells.append(f"{col}{r}={txt}")
        if cells:
            w("  " + " | ".join(cells))
    w()

# formula 시트: 부재코드×공종 산출식 (텍스트)
for nm in wb.sheetnames:
    low = nm.lower()
    if "formula" in low:
        dump_sheet(nm, max_r=80, max_c=15)
    elif "calc" in low:
        dump_sheet(nm, max_r=80, max_c=4, with_value=True)
    elif "csv" in low or "import" in low:
        dump_sheet(nm, max_r=8, max_c=16)
    elif "report" in low or "quantity" in low or "산출" in nm:
        dump_sheet(nm, max_r=40, max_c=27, with_value=True)

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print("WROTE", OUT, "lines", len(lines))
