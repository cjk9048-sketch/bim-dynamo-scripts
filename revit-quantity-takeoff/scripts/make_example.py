"""샘플 JSON 으로 예시 수량산출서를 생성하고 결과를 출력한다.

    py scripts/make_example.py
"""
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from takeoff import runner

src = ROOT / "samples" / "sample_members.json"
out = ROOT / "samples" / "수량산출서_예시.xlsx"
res = runner.run_from_json(str(src), str(out))

print("OUT:", res["out_path"])
print("부재수(인스턴스):", res["member_count"], "| 산출행:", res["row_count"])
print("합계:", json.dumps(res["totals"], ensure_ascii=False))

import openpyxl

wb = openpyxl.load_workbook(out)
print("시트:", wb.sheetnames)
ws = wb["콘크리트 (m³)"]
print("--- 콘크리트 시트 (No/공종/기호/산출식/수량/단위) ---")
for r in range(1, ws.max_row + 1):
    vals = [ws.cell(row=r, column=c).value for c in (1, 2, 3, 6, 7, 8)]
    print(r, vals)
