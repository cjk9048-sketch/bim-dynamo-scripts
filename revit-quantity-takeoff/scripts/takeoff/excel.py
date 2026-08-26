"""openpyxl 로 수량산출서 엑셀 파일을 만든다.

산출항목(콘크리트/거푸집/수량/길이)마다 시트를 하나씩 만들고,
카테고리별 소계와 전체 합계를 포함한다. 맨 앞에 '집계' 시트를 둔다.
"""
from __future__ import annotations

from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import CATEGORIES, CATEGORY_LABELS, LineItem

# (시트명, quantity 키, 숫자서식)
SHEET_GROUPS = [
    ("콘크리트 (m³)", "concrete", "#,##0.000"),
    ("거푸집 (m²)", "formwork", "#,##0.000"),
    ("부재수량 (EA)", "count", "#,##0"),
    ("부재길이 (m)", "length", "#,##0.000"),
]

HEADERS = ["No", "공종", "부재기호", "규격", "층", "산출식", "수량", "단위", "비고"]
_COL_WIDTHS = [5, 10, 10, 22, 8, 30, 12, 6, 26]

_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_SUBTOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
_TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")


def _style_row(ws, row: int, *, bold: bool = False, fill: PatternFill | None = None) -> None:
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = _BORDER
        if bold:
            cell.font = _BOLD
        if fill is not None:
            cell.fill = fill


def _cat_order(cat: str) -> int:
    return CATEGORIES.index(cat) if cat in CATEGORIES else len(CATEGORIES)


def _write_quantity_sheet(ws, items: list[LineItem], number_format: str) -> float:
    # 헤더
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = _COL_WIDTHS[col - 1]
    ws.freeze_panes = "A2"

    # 카테고리 -> 행들
    by_cat: dict[str, list[LineItem]] = defaultdict(list)
    for it in items:
        by_cat[it.category].append(it)

    row = 2
    no = 1
    grand_total = 0.0
    for cat in sorted(by_cat, key=_cat_order):
        cat_items = sorted(by_cat[cat], key=lambda x: (x.mark, x.type_name))
        subtotal = 0.0
        for it in cat_items:
            ws.cell(row=row, column=1, value=no).alignment = _CENTER
            ws.cell(row=row, column=2, value=CATEGORY_LABELS.get(it.category, it.category)).alignment = _CENTER
            ws.cell(row=row, column=3, value=it.mark).alignment = _CENTER
            ws.cell(row=row, column=4, value=it.type_name).alignment = _LEFT
            ws.cell(row=row, column=5, value=it.level).alignment = _CENTER
            ws.cell(row=row, column=6, value=it.formula).alignment = _LEFT
            vcell = ws.cell(row=row, column=7, value=it.value)
            vcell.number_format = number_format
            vcell.alignment = _RIGHT
            ws.cell(row=row, column=8, value=it.unit).alignment = _CENTER
            ws.cell(row=row, column=9, value=it.note).alignment = _LEFT
            _style_row(ws, row)
            subtotal += it.value
            grand_total += it.value
            no += 1
            row += 1

        # 카테고리 소계
        ws.cell(row=row, column=2, value=f"{CATEGORY_LABELS.get(cat, cat)} 소계")
        scell = ws.cell(row=row, column=7, value=round(subtotal, 3))
        scell.number_format = number_format
        scell.alignment = _RIGHT
        ws.cell(row=row, column=8, value=cat_items[0].unit if cat_items else "").alignment = _CENTER
        _style_row(ws, row, bold=True, fill=_SUBTOTAL_FILL)
        row += 1

    # 전체 합계
    ws.cell(row=row, column=2, value="합계")
    tcell = ws.cell(row=row, column=7, value=round(grand_total, 3))
    tcell.number_format = number_format
    tcell.alignment = _RIGHT
    ws.cell(row=row, column=8, value=items[0].unit if items else "").alignment = _CENTER
    _style_row(ws, row, bold=True, fill=_TOTAL_FILL)
    return round(grand_total, 3)


def _write_summary_sheet(ws, totals: dict[str, float], project_info: dict | None) -> None:
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 8
    r = 1
    title = ws.cell(row=r, column=1, value="수량 집계표")
    title.font = Font(bold=True, size=14)
    r += 2
    if project_info:
        for key, val in project_info.items():
            ws.cell(row=r, column=1, value=str(key)).font = _BOLD
            ws.cell(row=r, column=2, value=str(val))
            r += 1
        r += 1

    for col, name in enumerate(["산출항목", "합계", "단위"], start=1):
        cell = ws.cell(row=r, column=col, value=name)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    r += 1
    label_unit = {
        "concrete": ("콘크리트", "m³", "#,##0.000"),
        "formwork": ("거푸집", "m²", "#,##0.000"),
        "count": ("부재수량", "EA", "#,##0"),
        "length": ("부재길이", "m", "#,##0.000"),
    }
    for _, q, _fmt in SHEET_GROUPS:
        if q not in totals:
            continue
        label, unit, numfmt = label_unit[q]
        ws.cell(row=r, column=1, value=label).border = _BORDER
        vc = ws.cell(row=r, column=2, value=totals[q])
        vc.number_format = numfmt
        vc.alignment = _RIGHT
        vc.border = _BORDER
        uc = ws.cell(row=r, column=3, value=unit)
        uc.alignment = _CENTER
        uc.border = _BORDER
        r += 1


def write_takeoff(line_items: list[LineItem], out_path: str,
                  project_info: dict | None = None) -> dict[str, float]:
    """LineItem 리스트로 엑셀 수량산출서를 만들고 항목별 합계를 반환한다."""
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "집계"

    by_quantity: dict[str, list[LineItem]] = defaultdict(list)
    for it in line_items:
        by_quantity[it.quantity].append(it)

    totals: dict[str, float] = {}
    for sheet_name, q, numfmt in SHEET_GROUPS:
        items = by_quantity.get(q)
        if not items:
            continue
        ws = wb.create_sheet(title=sheet_name)
        totals[q] = _write_quantity_sheet(ws, items, numfmt)

    _write_summary_sheet(summary_ws, totals, project_info)

    wb.save(out_path)
    return totals
