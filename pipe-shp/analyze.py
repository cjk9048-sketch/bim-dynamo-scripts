# -*- coding: utf-8 -*-
import openpyxl, os
from collections import Counter

DATA = os.path.join(os.path.dirname(__file__), "data")
F1 = os.path.join(DATA, "DOM_SWR_v5_model_data.xlsx")
F2 = os.path.join(DATA, "Hydraulic_Results-(Min Slope 0.2%)_Formatted_v0.2.xlsx")

# --- File1 Nodes -> dict ---
wb1 = openpyxl.load_workbook(F1, read_only=True, data_only=True)
ws = wb1["Nodes"]
nodes = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    name, x, y = r[0], r[1], r[2]
    if name is None:
        continue
    nodes[str(name).strip()] = (x, y)
wb1.close()
print("File1 Nodes count:", len(nodes))

def prefix(s):
    s = str(s)
    p = ""
    for ch in s:
        if ch.isalpha() or ch in "-_":
            p += ch
        else:
            break
    return p or "(num)"

print("File1 node name prefixes:", Counter(prefix(k) for k in nodes).most_common())
samp = list(nodes.items())[:3]
print("File1 sample nodes:", samp)

# --- File2 Conduits ---
wb2 = openpyxl.load_workbook(F2, read_only=True, data_only=True)
cs = wb2["Conduit Output"]
rows = list(cs.iter_rows(min_row=2, values_only=True))
wb2.close()
print("\nFile2 Conduit rows:", len(rows))

starts = [str(r[1]).strip() if r[1] is not None else None for r in rows]
stops  = [str(r[2]).strip() if r[2] is not None else None for r in rows]
allref = [s for s in starts+stops if s]
print("File2 start/stop prefixes:", Counter(prefix(s) for s in allref).most_common())
print("File2 sample start/stop:", list(zip(starts[:3], stops[:3])))

missing = set()
matched = 0
for s in allref:
    if s in nodes:
        matched += 1
    else:
        missing.add(s)
print("\nRefs total:", len(allref), "matched:", matched, "missing:", len(allref)-matched)
print("Distinct missing node names:", len(missing))
print("Sample missing:", list(missing)[:10])

# rows where BOTH ends resolve
ok_rows = sum(1 for r in rows if str(r[1]).strip() in nodes and str(r[2]).strip() in nodes)
print("Conduit rows with BOTH ends found:", ok_rows, "/", len(rows))

# coordinate ranges
xs = [v[0] for v in nodes.values() if isinstance(v[0],(int,float))]
ys = [v[1] for v in nodes.values() if isinstance(v[1],(int,float))]
print("\nX range:", min(xs), "..", max(xs))
print("Y range:", min(ys), "..", max(ys))
