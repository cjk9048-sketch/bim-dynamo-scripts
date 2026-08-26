# -*- coding: utf-8 -*-
import openpyxl, os

DATA = os.path.join(os.path.dirname(__file__), "data")

def inspect(fn):
    path = os.path.join(DATA, fn)
    print("="*70)
    print("FILE:", fn)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print("SHEETS:", wb.sheetnames)
    for ws in wb.worksheets:
        print("-"*60)
        print("SHEET:", ws.title, "dims:", ws.max_row, "x", ws.max_column)
        rows = list(ws.iter_rows(min_row=1, max_row=6, values_only=True))
        for i, r in enumerate(rows, 1):
            # print up to 18 columns
            print(f"  r{i}:", [c for c in r[:18]])
    wb.close()

inspect("DOM_SWR_v5_model_data.xlsx")
inspect("Hydraulic_Results-(Min Slope 0.2%)_Formatted_v0.2.xlsx")
