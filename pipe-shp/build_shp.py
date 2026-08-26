# -*- coding: utf-8 -*-
"""
관로(선) Shapefile 생성기
- File1 Nodes 시트: Node 이름 -> (X, Y)
- File2 Conduit Output 시트: A=Label, B=Start Node, C=Stop Node, A~P=속성
- 각 관로를 시작노드->끝노드 직선으로 생성
"""
import openpyxl, os, csv
import shapefile  # pyshp

BASE = os.path.dirname(__file__)
DATA = os.path.join(BASE, "data")
OUT  = os.path.join(BASE, "output")
os.makedirs(OUT, exist_ok=True)

F1 = os.path.join(DATA, "DOM_SWR_v5_model_data.xlsx")
F2 = os.path.join(DATA, "Hydraulic_Results-(Min Slope 0.2%)_Formatted_v0.2.xlsx")
OUT_SHP = os.path.join(OUT, "pipes")

# 좌표 단위: 피트 -> 미터 변환
UNIT_SCALE = 0.3048
# 부여할 좌표계: WGS84 / UTM zone 36S (EPSG:32736) — 도도마(33°E존)
PRJ_WKT = (
    'PROJCS["WGS 84 / UTM zone 36S",'
    'GEOGCS["WGS 84",DATUM["WGS_1984",'
    'SPHEROID["WGS 84",6378137,298.257223563,AUTHORITY["EPSG","7030"]],'
    'AUTHORITY["EPSG","6326"]],'
    'PRIMEM["Greenwich",0,AUTHORITY["EPSG","8901"]],'
    'UNIT["degree",0.0174532925199433,AUTHORITY["EPSG","9122"]],'
    'AUTHORITY["EPSG","4326"]],'
    'PROJECTION["Transverse_Mercator"],'
    'PARAMETER["latitude_of_origin",0],'
    'PARAMETER["central_meridian",33],'
    'PARAMETER["scale_factor",0.9996],'
    'PARAMETER["false_easting",500000],'
    'PARAMETER["false_northing",10000000],'
    'UNIT["metre",1,AUTHORITY["EPSG","9001"]],'
    'AXIS["Easting",EAST],AXIS["Northing",NORTH],'
    'AUTHORITY["EPSG","32736"]]'
)

# ---- 1) 노드 사전 ----
wb1 = openpyxl.load_workbook(F1, read_only=True, data_only=True)
ws = wb1["Nodes"]
nodes = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    name, x, y = r[0], r[1], r[2]
    if name is None or x is None or y is None:
        continue
    nodes[str(name).strip()] = (float(x), float(y))
wb1.close()
print("노드 수:", len(nodes))

# ---- 2) 관로 읽기 ----
wb2 = openpyxl.load_workbook(F2, read_only=True, data_only=True)
cs = wb2["Conduit Output"]
all_rows = list(cs.iter_rows(min_row=1, values_only=True))
wb2.close()
header = all_rows[0]
rows = all_rows[1:]
print("관로 행 수:", len(rows))

# A~P = 16개 컬럼 (index 0..15)
N_ATTR = 16
orig_headers = [str(header[i]) if header[i] is not None else f"col{i+1}" for i in range(N_ATTR)]

# DBF 필드명(<=10자) 매핑
short_names = [
    "Label", "StartNode", "StopNode", "Diam_mm", "Flow_Ls", "Slope",
    "Length_m", "ManningN", "Vel_ms", "Cap_Ls", "InvStart_m", "InvStop_m",
    "Material", "DepthMid_m", "GrdStart_m", "GrdStop_m",
]
# 텍스트 컬럼 index: Label(0), StartNode(1), StopNode(2), Material(12)
TEXT_IDX = {0, 1, 2, 12}
INT_IDX = {3}  # Diameter

def safe_num(v):
    """숫자로 변환, 실패시(#N/A 등) None 반환"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "" or s.startswith("#"):
        return None
    try:
        return float(s)
    except ValueError:
        return None

# ---- 3) Shapefile Writer ----
w = shapefile.Writer(OUT_SHP, shapeType=shapefile.POLYLINE, encoding="utf-8")
for i in range(N_ATTR):
    nm = short_names[i]
    if i in TEXT_IDX:
        w.field(nm, "C", size=40)
    elif i in INT_IDX:
        w.field(nm, "N", size=10, decimal=0)
    else:
        w.field(nm, "N", size=19, decimal=6)

made = 0
dropped = []
for r in rows:
    label = r[0]
    s_name = str(r[1]).strip() if r[1] is not None else None
    e_name = str(r[2]).strip() if r[2] is not None else None
    if s_name not in nodes or e_name not in nodes:
        dropped.append((label, s_name, e_name))
        continue
    sx, sy = nodes[s_name]
    ex, ey = nodes[e_name]
    # 피트 -> 미터 변환
    sx, sy, ex, ey = sx*UNIT_SCALE, sy*UNIT_SCALE, ex*UNIT_SCALE, ey*UNIT_SCALE
    w.line([[[sx, sy], [ex, ey]]])
    attrs = []
    for i in range(N_ATTR):
        v = r[i] if i < len(r) else None
        if i in TEXT_IDX:
            attrs.append("" if v is None else str(v))
        else:
            attrs.append(safe_num(v))
    w.record(*attrs)
    made += 1

w.close()

# ---- 좌표계(.prj) 및 인코딩(.cpg) 파일 생성 ----
with open(OUT_SHP + ".prj", "w", encoding="utf-8") as f:
    f.write(PRJ_WKT)
with open(OUT_SHP + ".cpg", "w", encoding="ascii") as f:
    f.write("UTF-8")
print(".prj/.cpg 생성: EPSG:32736 (WGS84 / UTM 36S), 미터")

print("생성된 선:", made)
print("제외된 관로:", len(dropped))
for d in dropped:
    print("  제외:", d)

# ---- 4) 필드명 설명 범례 저장 ----
legend = os.path.join(OUT, "field_legend.csv")
with open(legend, "w", newline="", encoding="utf-8-sig") as f:
    wr = csv.writer(f)
    wr.writerow(["엑셀열", "SHP속성명", "원본 헤더(의미)"])
    for i in range(N_ATTR):
        col_letter = chr(ord("A") + i)
        wr.writerow([col_letter, short_names[i], orig_headers[i]])
print("범례 저장:", legend)
print("출력 폴더:", OUT)
